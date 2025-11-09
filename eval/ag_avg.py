#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
示例：
python json2xlsx.py \
  --folder data/jsons \
  --pattern "*.json" \
  --out_xlsx metrics/scores.xlsx \
  --overall_mode sample       # none | sample | category
"""

import os, json, glob, argparse
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- 配置 ----------
TARGET_CATEGORIES = [
    "Experimental Rigor",
    "Methodological Soundness",
    "Novelty & Significance",
    "Presentation & Clarity",
]
DIM_FIELDS = ["Attitude", "Clarity", "Persuasiveness", "Constructiveness"]
EXCLUDE_BASENAMES = ["test_data.json", "test_data1.json","test.json"]

# ---------- 工具 ----------
def extract_records(data):
    if isinstance(data, dict) and isinstance(data.get("data"), list):
        return data["data"]
    if isinstance(data, list):
        return data
    raise ValueError("Unsupported JSON root structure")

def safe_float(x):
    try:
        return float(x)
    except Exception:
        return None

# ---------- 统计 ----------
def compute_avgs(records, overall_mode="none"):
    """
    返回:
        cat_dim_avg  {category: {dim: avg}}
        overall_avg  float | None
        dim_overall  {dim: avg}       (4 维度总体均值)
    """
    buckets = {cat: {f: [] for f in DIM_FIELDS} for cat in TARGET_CATEGORIES}
    dim_pool = {f: [] for f in DIM_FIELDS}

    for r in records:
        cat = r.get("comment", {}).get("category")
        if cat not in TARGET_CATEGORIES:
            continue
        scores = r.get("refine", {}).get("score", {})
        vals = [safe_float(scores.get(f)) for f in DIM_FIELDS]
        if any(v is None for v in vals):          # 任一维度缺失则忽略此条
            continue
        for f, v in zip(DIM_FIELDS, vals):
            buckets[cat][f].append(v)
            dim_pool[f].append(v)

    cat_dim_avg = {
        cat: {f: (round(mean(buckets[cat][f]), 3) if buckets[cat][f] else None)
              for f in DIM_FIELDS}
        for cat in TARGET_CATEGORIES
    }

    dim_overall = {
        f: (round(mean(dim_pool[f]), 3) if dim_pool[f] else None)
        for f in DIM_FIELDS
    }

    overall_avg = None
    if overall_mode == "sample":
        all_vals = [v for vs in dim_pool.values() for v in vs]
        overall_avg = round(mean(all_vals), 3) if all_vals else None
    elif overall_mode == "category":
        cat_means = [mean([v for v in cat_dim_avg[c].values() if v is not None])
                     for c in TARGET_CATEGORIES
                     if any(cat_dim_avg[c].values())]
        overall_avg = round(mean(cat_means), 3) if cat_means else None
    elif overall_mode != "none":
        raise ValueError("overall_mode must be none|sample|category")

    return cat_dim_avg, overall_avg, dim_overall

# ---------- 主流程 ----------
def main(folder, pattern, out_xlsx, overall_mode):
    files = sorted(glob.glob(os.path.join(folder, pattern)))
    files = [f for f in files if os.path.basename(f) not in EXCLUDE_BASENAMES]
    if not files:
        print("No files found.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # ===== 表头 =====
    ws.cell(row=1, column=1, value="")
    ws.cell(row=2, column=1, value="file").alignment = center
    col = 2
    for cat in TARGET_CATEGORIES:
        start = col
        for f in DIM_FIELDS:
            ws.cell(row=2, column=col, value=f).alignment = center
            col += 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col-1)
        ws.cell(row=1, column=start, value=cat).alignment = center

    # ---- 追加 4 列维度总体均值 ----
    dim_start = col
    for f in DIM_FIELDS:
        ws.cell(row=2, column=col, value=f + "_mean").alignment = center
        col += 1
    ws.merge_cells(start_row=1, start_column=dim_start, end_row=1, end_column=col-1)
    ws.cell(row=1, column=dim_start, value="Dim Overall").alignment = center

    # ---- overall_avg 列（可选）----
    overall_col = None
    if overall_mode != "none":
        overall_col = col
        ws.cell(row=1, column=overall_col, value="overall_avg").alignment = center
        ws.cell(row=2, column=overall_col, value=overall_mode).alignment = center
        col += 1

    # ===== 数据行 =====
    row_idx = 3
    for fp in files:
        fn = os.path.splitext(os.path.basename(fp))[0]
        ws.cell(row=row_idx, column=1, value=fn)

        with open(fp, encoding="utf-8") as f:
            records = extract_records(json.load(f))

        cat_dim_avg, overall_avg, dim_overall = compute_avgs(records, overall_mode)

        c = 2
        for cat in TARGET_CATEGORIES:
            for f in DIM_FIELDS:
                ws.cell(row=row_idx, column=c, value=cat_dim_avg[cat][f]); c += 1
        for f in DIM_FIELDS:
            ws.cell(row=row_idx, column=c, value=dim_overall[f]); c += 1
        if overall_col:
            ws.cell(row=row_idx, column=overall_col, value=overall_avg)
        row_idx += 1

    # ===== 样式 =====
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    wb.save(out_xlsx)
    print(f"Saved to {out_xlsx}")

# ---------- CLI ----------
if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--folder", required=True, help="包含 JSON 的文件夹")
    p.add_argument("--pattern", default="*.json", help="glob 匹配模式")
    p.add_argument("--out_xlsx", default="metrics/new.xlsx")
    p.add_argument("--overall_mode",
                   choices=["none", "sample", "category"],
                   default="none",
                   help="overall_avg 计算方式")
    main(**vars(p.parse_args()))